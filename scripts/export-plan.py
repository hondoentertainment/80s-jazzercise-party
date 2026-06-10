from __future__ import annotations

import json
from datetime import datetime, time
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "Project Plan for the 41st Birthday Party.xlsx"
OUTPUT = ROOT / "js" / "plan-data.js"


def format_time(value) -> str:
    if value is None or value == "TBD":
        return "TBD"
    if isinstance(value, time):
        hour = value.hour % 12 or 12
        minute = value.minute
        suffix = "AM" if value.hour < 12 else "PM"
        if minute:
            return f"{hour}:{minute:02d} {suffix}"
        return f"{hour} {suffix}"
    text = str(value).strip()
    if not text:
        return "TBD"
    return text


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def export_tasks(ws) -> list[dict]:
    tasks = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        try:
            task_id = int(row[0])
        except (TypeError, ValueError):
            continue
        tasks.append(
            {
                "id": task_id,
                "category": clean(row[1]),
                "task": clean(row[2]),
                "status": clean(row[3]) or "Not Started",
                "assigned": clean(row[4]),
                "notes": clean(row[5]) if len(row) > 5 else "",
            }
        )
    return tasks


def export_schedule(ws) -> list[dict]:
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not any(row):
            continue
        day = clean(row[0])
        action = clean(row[2]) if len(row) > 2 else ""
        if not day or not action:
            continue
        rows.append(
            {
                "day": day,
                "time": format_time(row[1] if len(row) > 1 else "TBD"),
                "action": action,
            }
        )
    return rows


def export_signs(ws) -> list[dict]:
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name = clean(row[0])
        if name.lower() == "sign list":
            continue
        quantity = row[1]
        try:
            qty = int(quantity)
        except (TypeError, ValueError):
            qty = clean(quantity) or "1"
        rows.append({"name": name, "quantity": qty})
    return rows


def export_committee(ws) -> list[dict]:
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        name = clean(row[0])
        role = clean(row[1]) if len(row) > 1 else ""
        if not name:
            continue
        rows.append({"name": name, "role": role})
    return rows


def normalize_plan(plan: dict) -> dict:
    title = plan.get("title", "")
    if not title or "Eighties" in title or "Jazzercise" in title or title.endswith("TBD"):
        plan["title"] = "Cowboy Disco Party — August 15, 2026"
    plan["subtitle"] = "Project plan for the 41st birthday party — Cowboy Disco theme"
    auto_complete = {
        "Pick Theme": ("Complete", "Cowboy Disco — boots, bling, fringe & sequins"),
        "Design Party Website": ("Complete", "cowboy-disco-party.vercel.app"),
        "Design Ice Breaker": ("Complete", "15-card deck at /ice-breaker.html"),
        "Design Food Signs": ("Complete", "Saloon fuel labels in /signs.html"),
        "Print Food Signs": ("Complete", "Print pack at /print-pack.html"),
        "Design Team Game": ("Complete", "Game design finished"),
        "Pick Date": ("Complete", "Saturday, August 15, 2026"),
        "Pick Out Menu": ("Complete", "Menu PDF + drink list on site"),
        "Design Signs": ("Complete", "Entrance + Saloon signs in /signs.html"),
        "Print Signs": ("Complete", "/print-pack.html"),
        "Centralized Photo Share": ("Complete", "Gallery at /gallery.html"),
        "Design Partiful Invite": ("Complete", "Digital invite at /invite.html"),
        "Send Partiful Invite": ("In Progress", "Use /invite.html — copy, SMS, or email"),
    }
    for task in plan.get("tasks", []):
        name = task.get("task", "")
        if name in auto_complete:
            task["status"], task["notes"] = auto_complete[name]
    for sign in plan.get("signs", []):
        if sign.get("name") == "Kyle's Apartment":
            sign["name"] = "Cowboy Disco Saloon"
    return plan


def main() -> None:
    wb = load_workbook(WORKBOOK, data_only=True)
    plan = normalize_plan(
        {
            "title": clean(wb["Notes"].cell(1, 1).value) or "Cowboy Disco Party TBD",
            "subtitle": "Project plan for the 41st birthday party — Cowboy Disco theme",
            "sourceFile": WORKBOOK.name,
            "exportedAt": datetime.now().isoformat(timespec="seconds"),
            "tasks": export_tasks(wb["Notes"]),
            "schedule": export_schedule(wb["Schedule"]),
            "signs": export_signs(wb["Signs"]),
            "committee": export_committee(wb["Committe"]),
        }
    )

    payload = json.dumps(plan, indent=2, ensure_ascii=False)
    OUTPUT.write_text(
        "(function (global) {\n"
        '  "use strict";\n\n'
        f"  global.CDP_PLAN = {payload};\n"
        "})(window);\n",
        encoding="utf-8",
    )
    print(f"Wrote {OUTPUT.name} ({len(plan['tasks'])} tasks)")


if __name__ == "__main__":
    main()
